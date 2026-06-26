"""Generacion de certificados PDF a partir de un Lote (excel + plantilla)."""

import io
import re
import shutil
import unicodedata
import zipfile
from pathlib import Path

import fitz
import openpyxl
from PIL import Image
from reportlab.lib.colors import HexColor
from reportlab.pdfbase.pdfmetrics import stringWidth
from reportlab.pdfgen import canvas


class GeneradorError(Exception):
    pass


def limpiar_nombre(valor) -> str | None:
    if valor is None:
        return None
    texto = str(valor).strip()
    texto = re.sub(r"\s+", " ", texto)
    texto = texto.strip(" ,.-:;")
    return texto or None


def nombre_archivo_seguro(nombre: str) -> str:
    sin_acentos = unicodedata.normalize("NFKD", nombre).encode("ascii", "ignore").decode("ascii")
    limpio = re.sub(r"[^A-Za-z0-9 _-]", "", sin_acentos).strip()
    limpio = re.sub(r"\s+", "_", limpio)
    return limpio or "certificado"


def leer_nombres(ruta_excel) -> list[str]:
    wb = openpyxl.load_workbook(ruta_excel, data_only=True)
    ws = wb.active

    nombres = []
    vistos = set()
    for fila in ws.iter_rows(min_row=1, max_col=1):
        nombre = limpiar_nombre(fila[0].value)
        if nombre is None:
            continue
        clave = nombre.lower()
        if clave in vistos:
            continue
        vistos.add(clave)
        nombres.append(nombre)
    return nombres


PATRON_CARACTERES_RAROS = re.compile(r"[^A-Za-zÁÉÍÓÚÑÜáéíóúñü\s'.-]")
PATRON_SOLO_LETRAS = re.compile(r"[A-Za-zÁÉÍÓÚÑÜáéíóúñü]")


def analizar_excel(ruta_excel) -> dict:
    """Recorre el Excel fila por fila y devuelve la lista final de nombres
    junto con advertencias sobre filas problematicas (vacias, muy cortas,
    con caracteres raros o duplicadas), para que el usuario las revise
    antes de generar certificados.
    """
    wb = openpyxl.load_workbook(ruta_excel, data_only=True)
    ws = wb.active

    filas = []
    vistos = {}
    fila_excel = 0
    for fila in ws.iter_rows(min_row=1, max_col=1):
        fila_excel += 1
        valor_original = fila[0].value
        nombre = limpiar_nombre(valor_original)

        advertencias = []
        if nombre is None:
            if valor_original is not None and str(valor_original).strip() != "":
                advertencias.append("Fila vacia tras limpiar el texto")
            else:
                continue  # fila realmente vacia, no se reporta
        else:
            if len(nombre) < 4:
                advertencias.append("Nombre muy corto, revisar si esta incompleto")
            if not PATRON_SOLO_LETRAS.search(nombre):
                advertencias.append("No contiene letras")
            caracteres_raros = sorted(set(PATRON_CARACTERES_RAROS.findall(nombre)))
            if caracteres_raros:
                advertencias.append(f"Contiene caracteres inusuales: {' '.join(caracteres_raros)}")
            clave = nombre.lower()
            if clave in vistos:
                advertencias.append(f"Duplicado de la fila {vistos[clave]}")
            else:
                vistos[clave] = fila_excel

        filas.append({
            "fila": fila_excel,
            "valor_original": "" if valor_original is None else str(valor_original),
            "nombre": nombre,
            "advertencias": advertencias,
        })

    nombres_finales = leer_nombres(ruta_excel)
    return {
        "filas": filas,
        "total_filas_con_datos": len(filas),
        "total_nombres_validos": len(nombres_finales),
        "filas_con_advertencias": [f for f in filas if f["advertencias"]],
    }


def exportar_reporte_excel(analisis: dict) -> bytes:
    """Genera un Excel descargable con el detalle de filas y advertencias de analizar_excel."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Reporte"

    encabezados = ["Fila", "Valor original", "Nombre que se usara", "Advertencias"]
    ws.append(encabezados)
    for celda in ws[1]:
        celda.font = openpyxl.styles.Font(bold=True)

    relleno_advertencia = openpyxl.styles.PatternFill(start_color="FFF3CD", end_color="FFF3CD", fill_type="solid")

    for item in analisis["filas"]:
        fila_valores = [
            item["fila"],
            item["valor_original"],
            item["nombre"] or "(vacio)",
            "; ".join(item["advertencias"]),
        ]
        ws.append(fila_valores)
        if item["advertencias"]:
            for celda in ws[ws.max_row]:
                celda.fill = relleno_advertencia

    for columna, ancho in zip("ABCD", (8, 35, 35, 50)):
        ws.column_dimensions[columna].width = ancho

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer.read()


def calcular_tamano_fuente(texto: str, fuente: str, tamano_inicial: int, tamano_min: int, ancho_max_pt: float) -> int:
    tamano = tamano_inicial
    while tamano > tamano_min and stringWidth(texto, fuente, tamano) > ancho_max_pt:
        tamano -= 1
    return tamano


def preparar_plantilla_comprimida(template_path: Path, carpeta_temporal: Path, calidad: int = 85) -> Path:
    """Convierte la plantilla a JPEG calidad alta una sola vez, para que cada
    PDF generado embeba esa version liviana en vez de la imagen original sin
    comprimir. Reduce el peso de cada certificado entre 5x y 8x.
    """
    carpeta_temporal.mkdir(parents=True, exist_ok=True)
    destino = carpeta_temporal / "plantilla_comprimida.jpg"
    with Image.open(template_path) as img:
        img.convert("RGB").save(destino, "JPEG", quality=calidad, optimize=True)
    return destino


def _dibujar_certificado(c: canvas.Canvas, nombre: str, template_path: Path, ancho_pt: int, alto_pt: int, opciones: dict):
    c.drawImage(str(template_path), 0, 0, width=ancho_pt, height=alto_pt)

    fuente = opciones.get("fuente", "Helvetica-Bold")
    color = opciones.get("color", "#1a1a1a")
    tamano_max = opciones.get("tamano_fuente", 40)
    tamano_min = opciones.get("tamano_fuente_min", 18)

    x_px = opciones["x"]
    y_px = opciones["y"]
    y_pdf = alto_pt - y_px

    margen_izq = opciones.get("margen_izquierdo", 0)
    margen_der = opciones.get("margen_derecho", 0)
    espacio_izq = x_px - margen_izq
    espacio_der = (ancho_pt - margen_der) - x_px
    ancho_max = 2 * min(espacio_izq, espacio_der)

    tamano = calcular_tamano_fuente(nombre, fuente, tamano_max, tamano_min, max(ancho_max, 1))

    c.setFont(fuente, tamano)
    c.setFillColor(HexColor(color))
    c.drawCentredString(x_px, y_pdf, nombre)


def generar_pdf_certificado(nombre: str, pdf_path: Path, template_path: Path, img_w_px: int, img_h_px: int, opciones: dict):
    c = canvas.Canvas(str(pdf_path), pagesize=(img_w_px, img_h_px))
    _dibujar_certificado(c, nombre, template_path, img_w_px, img_h_px, opciones)
    c.showPage()
    c.save()


def opciones_desde_lote(lote) -> dict:
    return {
        "x": lote.texto_x,
        "y": lote.texto_y,
        "margen_izquierdo": lote.margen_izquierdo,
        "margen_derecho": lote.margen_derecho,
        "fuente": lote.fuente,
        "tamano_fuente": lote.tamano_fuente,
        "tamano_fuente_min": lote.tamano_fuente_min,
        "color": lote.color_texto,
    }


def generar_previsualizacion_png(lote, ancho_max_px: int = 1000) -> bytes:
    """Renderiza el certificado con un nombre de muestra y devuelve un PNG en memoria."""
    template_path = Path(lote.plantilla.path)
    if not template_path.exists():
        raise GeneradorError("No se encontro el archivo de plantilla.")

    with Image.open(template_path) as img:
        img_w_px, img_h_px = img.size

    nombre_muestra = "Nombre de Ejemplo Apellido"
    try:
        excel_path = Path(lote.excel.path)
        if excel_path.exists():
            nombres = leer_nombres(excel_path)
            if nombres:
                nombre_muestra = nombres[0]
    except Exception:
        pass

    buffer_pdf = io.BytesIO()
    c = canvas.Canvas(buffer_pdf, pagesize=(img_w_px, img_h_px))
    _dibujar_certificado(c, nombre_muestra, template_path, img_w_px, img_h_px, opciones_desde_lote(lote))
    c.showPage()
    c.save()
    buffer_pdf.seek(0)

    doc = fitz.open(stream=buffer_pdf.read(), filetype="pdf")
    zoom = min(ancho_max_px / img_w_px, 2.0)
    pix = doc[0].get_pixmap(matrix=fitz.Matrix(zoom, zoom))
    return pix.tobytes("png")


def contar_nombres(lote) -> int:
    excel_path = Path(lote.excel.path)
    if not excel_path.exists():
        raise GeneradorError("No se encontro el archivo Excel.")
    return len(leer_nombres(excel_path))


def generar_certificados_lote(lote, carpeta_salida: Path, on_progreso=None) -> tuple[int, Path]:
    """Genera todos los PDF del lote y un zip con el resultado.

    on_progreso(procesados, total) se llama tras cada PDF generado, para
    poder reportar avance en tiempo real desde quien llame a esta funcion.

    Devuelve (cantidad_generados, ruta_zip).
    """
    template_path = Path(lote.plantilla.path)
    excel_path = Path(lote.excel.path)

    if not template_path.exists():
        raise GeneradorError("No se encontro el archivo de plantilla.")
    if not excel_path.exists():
        raise GeneradorError("No se encontro el archivo Excel.")

    with Image.open(template_path) as img:
        img_w_px, img_h_px = img.size

    nombres = leer_nombres(excel_path)
    if not nombres:
        raise GeneradorError("El Excel no contiene nombres en la columna A.")

    carpeta_salida.mkdir(parents=True, exist_ok=True)

    carpeta_temporal = carpeta_salida.parent / "_tmp"
    plantilla_para_pdf = preparar_plantilla_comprimida(template_path, carpeta_temporal)

    opciones = opciones_desde_lote(lote)

    total = len(nombres)
    pdfs = []
    usados = {}
    try:
        for indice, nombre in enumerate(nombres, start=1):
            base = nombre_archivo_seguro(nombre)
            usados[base] = usados.get(base, 0) + 1
            sufijo = "" if usados[base] == 1 else f"_{usados[base]}"
            pdf_path = carpeta_salida / f"{base}{sufijo}.pdf"
            generar_pdf_certificado(nombre, pdf_path, plantilla_para_pdf, img_w_px, img_h_px, opciones)
            pdfs.append(pdf_path)
            if on_progreso:
                on_progreso(indice, total)

        zip_path = carpeta_salida / "certificados.zip"
        with zipfile.ZipFile(zip_path, "w", zipfile.ZIP_DEFLATED) as zf:
            for pdf_path in pdfs:
                zf.write(pdf_path, arcname=pdf_path.name)
    finally:
        shutil.rmtree(carpeta_temporal, ignore_errors=True)

    return len(pdfs), zip_path
